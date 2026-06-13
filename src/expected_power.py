"""Expected-power model (XGBoost) - an INDEPENDENT loss cross-check.

This is a backend cross-check ONLY. It is NOT in the agent decision path and it
NEVER raises an alarm - it merely re-prices an already-detected fault window with
a weather-driven counterfactual, alongside the Slice-2 sibling/CausalImpact one.

Per IEA PVPS Task 13 the model is trained ONLY on CLEAN points (night, low
irradiance, curtailment, clipping and outage/zero rows removed) so it learns
HEALTHY operation; predicting that healthy expected power over the dead window
and integrating the positive shortfall gives a second, independent lost-energy
estimate. See docs/RESEARCH_NOTES.md ("Expected-power model").
"""

from __future__ import annotations

import re

import duckdb
import numpy as np
import pandas as pd
from pydantic import BaseModel
from xgboost import XGBRegressor

from src.ingest import DEFAULT_SYSTEM_OVERVIEW, canonical_inverter_id, load_meta

NATIVE = "data/Plant A (start here)/1. Main-monitoring-data/main_monitoring_data.parquet"
FEATURES = ["irradiance", "module_temp", "ambient_temp", "sun_altitude"]
STEP_HOURS = 300 / 3600.0
MIN_IRRADIANCE = 50.0
CLIP_FRAC = 0.98  # P_AC >= 98% of rated kW = inverter clipping/saturation
YEAR1 = (2017, "2017-01-01", "2017-12-31")


class ExpectedPowerResult(BaseModel):
    """Validation metrics + explainability for one inverter's clean-trained model."""

    inverter_id: str
    year1_window: str
    val_mae: float
    val_rmse: float
    val_mbe: float
    val_r2: float
    val_pct_capacity: float  # RMSE as % of rated kWp
    feature_importances: dict
    shap_available: bool = False


def _rated(inverter_id: str) -> float:
    meta = load_meta()
    row = meta[meta["inverter_id"] == inverter_id]
    return float(row["kwp"].iloc[0]) if not row.empty else 0.0


def _load_series(inverter_id: str, start: str, end: str) -> pd.DataFrame:
    """Native-resolution feature + P_AC series for one inverter over [start, end]."""
    con = duckdb.connect()
    try:
        cols = [r[0] for r in con.execute(
            f"DESCRIBE SELECT * FROM read_parquet('{NATIVE}')").fetchall()]

        def find(p):
            return next((c for c in cols if re.search(p, c)), None)

        eid = re.escape(inverter_id)
        cmap = {
            "irradiance": find(r"Irradiation"),
            "module_temp": find(r"Temperature Sensor / Module"),
            "ambient_temp": find(r"Temperature Sensor / Ambient"),
            "sun_altitude": find(r"Altitude"),
            "dv": find(r"DRD11A / DV"),
            "p_ac": find(eid + r" / P_AC"),
        }
        sel = ", ".join(f'"{c}" AS {k}' for k, c in cmap.items())
        df = con.execute(
            f"""SELECT strptime(timestamp, '%Y.%m.%d %H:%M') AS ts, {sel}
                FROM read_parquet('{NATIVE}')
                WHERE CAST(strptime(timestamp, '%Y.%m.%d %H:%M') AS DATE)
                      BETWEEN '{start}' AND '{end}'
                ORDER BY ts"""
        ).df()
    finally:
        con.close()
    return df


def _clean_mask(df: pd.DataFrame, rated: float) -> pd.Series:
    """IEA Task 13 clean filter: producing daytime, not curtailed/clipped/dead."""
    return (
        (df["sun_altitude"] > 0)
        & (df["irradiance"] >= MIN_IRRADIANCE)
        & (df["dv"].fillna(100) >= 100)  # DV null = no signal = not curtailed
        & (df["p_ac"] > 0.1)
        & (df["p_ac"] < CLIP_FRAC * rated)
        & df[FEATURES + ["p_ac"]].notna().all(axis=1)
    )


def clean_training_frame(inverter_id: str, start: str, end: str):
    """Return (clean_df, rated) for inspection/tests."""
    iid = canonical_inverter_id(inverter_id) or inverter_id
    rated = _rated(iid)
    df = _load_series(iid, start, end)
    return df[_clean_mask(df, rated)].sort_values("ts").reset_index(drop=True), rated


def train_expected_power(inverter_id, year1_start=YEAR1[1], year1_end=YEAR1[2]):
    """Train an XGBoost expected-power model on year-1 CLEAN points; validate on a
    time-held-out clean split. Returns (model, ExpectedPowerResult)."""
    iid = canonical_inverter_id(inverter_id) or inverter_id
    clean, rated = clean_training_frame(iid, year1_start, year1_end)
    if len(clean) < 1000 or rated <= 0:
        raise ValueError(f"insufficient clean data for {iid} ({len(clean)} rows)")
    k = int(len(clean) * 0.8)
    tr, va = clean.iloc[:k], clean.iloc[k:]
    model = XGBRegressor(
        n_estimators=600, max_depth=5, learning_rate=0.05,
        subsample=0.8, colsample_bytree=0.8, min_child_weight=5,
        importance_type="gain", early_stopping_rounds=30,
        n_jobs=4, random_state=0,
    )
    model.fit(tr[FEATURES], tr["p_ac"], eval_set=[(va[FEATURES], va["p_ac"])], verbose=False)
    pred = np.asarray(model.predict(va[FEATURES]), dtype=float)
    y = va["p_ac"].to_numpy(dtype=float)
    err = pred - y
    ss_res = float(np.sum((y - pred) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
    rmse = float(np.sqrt(np.mean(err ** 2)))
    fi = {f: float(v) for f, v in zip(FEATURES, model.feature_importances_)}
    result = ExpectedPowerResult(
        inverter_id=iid, year1_window=f"{year1_start}..{year1_end}",
        val_mae=round(float(np.mean(np.abs(err))), 4),
        val_rmse=round(rmse, 4),
        val_mbe=round(float(np.mean(err)), 4),
        val_r2=round(r2, 4),
        val_pct_capacity=round(rmse / rated * 100.0, 3),
        feature_importances={f: round(v, 4) for f, v in fi.items()},
        shap_available=_shap_ok(model, tr[FEATURES]),
    )
    return model, result


def _shap_ok(model, x_sample) -> bool:
    """Confirm SHAP runs on the tree model (explainability mitigation); else False."""
    try:
        import shap

        shap.TreeExplainer(model).shap_values(x_sample.head(100))
        return True
    except Exception:
        return False


def predict_expected(model, inverter_id, window_start, window_end) -> pd.DataFrame:
    """DataFrame[ts, expected_p_ac, actual_p_ac, residual] over the window."""
    iid = canonical_inverter_id(inverter_id) or inverter_id
    df = _load_series(iid, window_start, window_end)
    feat = df.dropna(subset=FEATURES).copy()
    feat["expected_p_ac"] = np.clip(model.predict(feat[FEATURES]), 0, None)
    feat["actual_p_ac"] = feat["p_ac"]
    feat["residual"] = feat["actual_p_ac"] - feat["expected_p_ac"]
    feat["irradiance"] = df.loc[feat.index, "irradiance"]
    feat["sun_altitude"] = df.loc[feat.index, "sun_altitude"]
    feat["dv"] = df.loc[feat.index, "dv"]
    return feat[["ts", "expected_p_ac", "actual_p_ac", "residual", "irradiance",
                 "sun_altitude", "dv"]]


def xgb_lost_kwh(model, inverter_id, window_start, window_end) -> float:
    """Positive shortfall (expected-actual) integrated over clean-weather window intervals."""
    pe = predict_expected(model, inverter_id, window_start, window_end)
    producing = pe[(pe["sun_altitude"] > 0) & (pe["irradiance"] >= MIN_IRRADIANCE)
                   & (pe["dv"].fillna(100) >= 100)]
    shortfall = (producing["expected_p_ac"] - producing["actual_p_ac"]).clip(lower=0)
    return float(shortfall.sum() * STEP_HOURS)


def degradation_series(model, inverter_id, years=range(2017, 2026)) -> list:
    """Annual mean clean-conditions residual as % of rated capacity (a PLR proxy)."""
    iid = canonical_inverter_id(inverter_id) or inverter_id
    rated = _rated(iid)
    out = []
    for yr in years:
        df = _load_series(iid, f"{yr}-01-01", f"{yr}-12-31")
        clean = df[_clean_mask(df, rated)]
        if clean.empty:
            continue
        pred = np.clip(model.predict(clean[FEATURES]), 0, None)
        resid_pct = float(np.mean(clean["p_ac"].to_numpy() - pred) / rated * 100.0)
        out.append({"year": int(yr), "mean_residual_pct": round(resid_pct, 3)})
    return out


def plr_proxy_pct_per_year(series: list) -> float:
    """Linear slope of mean_residual_pct vs year (negative = degrading)."""
    if len(series) < 2:
        return 0.0
    yrs = np.array([s["year"] for s in series], dtype=float)
    val = np.array([s["mean_residual_pct"] for s in series], dtype=float)
    slope = float(np.polyfit(yrs, val, 1)[0])
    return round(slope, 4)


def _module_types() -> dict:
    """Map canonical inverter id -> module type string from System_Overview."""
    import openpyxl

    wb = openpyxl.load_workbook(DEFAULT_SYSTEM_OVERVIEW, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i = next(i for i, r in enumerate(rows)
                 if any(str(c).strip() == "Description" for c in r if c is not None))
    hdr = [("" if c is None else str(c).strip()) for c in rows[hdr_i]]
    desc_c = hdr.index("Description")
    mt_c = next((i for i, h in enumerate(hdr) if "Module Type" in h), None)
    out: dict[str, str] = {}
    for r in rows[hdr_i + 1:]:
        if not r or mt_c is None or len(r) <= mt_c:
            continue
        iid = canonical_inverter_id(r[desc_c])
        if iid and iid not in out and r[mt_c] not in (None, "", "-"):
            out[iid] = str(r[mt_c]).strip()
    return out


def benchmark(inverter_ids) -> list:
    """Train+validate each inverter; report val_r2, error% and a PLR proxy."""
    types = _module_types()
    rows = []
    for iid in inverter_ids:
        cid = canonical_inverter_id(iid) or iid
        model, res = train_expected_power(cid)
        ser = degradation_series(model, cid)
        rows.append({
            "inverter_id": cid,
            "module_type": types.get(cid, "unknown"),
            "val_r2": res.val_r2,
            "val_pct_capacity": res.val_pct_capacity,
            "plr_proxy_pct_per_year": plr_proxy_pct_per_year(ser),
        })
    return rows
