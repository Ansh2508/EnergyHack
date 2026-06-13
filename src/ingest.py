"""Ingest layer: monitoring parquet cache + plant metadata (DuckDB, off-disk).

Monitoring source of truth is the Plant A native parquet
(`main_monitoring_data.parquet`) -- already typed and clean. The German CSV
(`main_monitoring_data.csv`) is byte-corrupted around line 660499
(non-UTF-8 sequence), so the parquet is preferred and the CSV is supported only
as a resilient fallback (strict_mode off, ignore_errors). Content is identical
across formats per the plant's "General information" sheet.

Confirmed parse facts (see outputs/schema_verification.txt): 990,442 rows,
2016-12-31 -> 2026-06-01, modal step 300 s, 65 inverters, timestamp format
'%Y.%m.%d %H:%M', decimal ',', delimiter ';'.
"""

from __future__ import annotations

import os
import re

import duckdb
import pandas as pd

STEP_SECONDS = 300

_DATA = "data/Plant A (start here)"
DEFAULT_MONITORING = f"{_DATA}/1. Main-monitoring-data/main_monitoring_data.parquet"
DEFAULT_CSV = f"{_DATA}/1. Main-monitoring-data/main_monitoring_data.csv"
DEFAULT_CACHE = "data/_cache/monitoring.parquet"
DEFAULT_SYSTEM_OVERVIEW = f"{_DATA}/2. Additional Data/System_Overview.xlsx"

_INV_RE = re.compile(r"INV \d+\.\d+\.\d+")
_TS_FMT = "%Y.%m.%d %H:%M"


def canonical_inverter_id(text: str) -> str | None:
    """Return canonical 'INV bb.ss.nnn' from a column name or meta label.

    Handles monitoring columns ('INV 01.05.029 / P_AC (kW)') and System_Overview
    descriptions ('WR 01 .05 .029' -> 'INV 01.05.029'). Split sub-inverters
    ('WR 01 .01. 004.02') collapse onto their 3-part parent ('INV 01.01.004').
    """
    m = _INV_RE.search(str(text))
    if m:
        return m.group(0)
    nums = re.findall(r"\d+", str(text))
    return f"INV {nums[0]}.{nums[1]}.{nums[2]}" if len(nums) >= 3 else None


def resolve_columns(cols: list[str]) -> dict:
    """Resolve monitoring column groups by regex -- never hardcoded strings."""

    def first(pat: str) -> str | None:
        return next((c for c in cols if re.search(pat, c, re.I)), None)

    return {
        "pac": [c for c in cols if re.search(r"INV \d+\.\d+\.\d+ / P_AC", c)],
        "idc": [c for c in cols if re.search(r"INV \d+\.\d+\.\d+ / I_DC", c)],
        "udc": [c for c in cols if re.search(r"INV \d+\.\d+\.\d+ / U_DC", c)],
        "irr": first(r"Irradiation"),
        "alt": first(r"Altitude"),
        "tmod": first(r"/ Module"),
        "dv": first(r"/ DV"),
    }


def build_monitoring_cache(source_path: str, cache_path: str) -> str:
    """Build the typed monitoring parquet cache once (idempotent).

    Prefers a native parquet source; falls back to the German CSV. Only the
    columns Slice 1 needs are cached (timestamp + P_AC/I_DC/U_DC + env tracks).
    """
    if os.path.exists(cache_path):
        return cache_path
    parent = os.path.dirname(cache_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    is_parquet = source_path.lower().endswith(".parquet")
    reader = (
        f"read_parquet('{source_path}')"
        if is_parquet
        else (
            f"read_csv('{source_path}', delim=';', header=true, all_varchar=true, "
            f"strict_mode=false, ignore_errors=true)"
        )
    )
    con = duckdb.connect()
    cols = [r[0] for r in con.execute(f"DESCRIBE SELECT * FROM {reader}").fetchall()]
    g = resolve_columns(cols)
    keep = (
        g["pac"]
        + g["idc"]
        + g["udc"]
        + [c for c in (g["irr"], g["alt"], g["tmod"], g["dv"]) if c]
    )
    print(
        f"[ingest] resolved PAC={len(g['pac'])} IDC={len(g['idc'])} "
        f"UDC={len(g['udc'])} irr={g['irr']!r} alt={g['alt']!r} dv={g['dv']!r}"
    )
    ts = f"strptime(\"timestamp\", '{_TS_FMT}') AS \"timestamp\""
    if is_parquet:
        body = [ts] + [f'"{c}"' for c in keep]  # native parquet is already DOUBLE
    else:
        body = [ts] + [
            f"TRY_CAST(replace(\"{c}\", ',', '.') AS DOUBLE) AS \"{c}\"" for c in keep
        ]
    con.execute(
        f"COPY (SELECT {', '.join(body)} FROM {reader}) "
        f"TO '{cache_path}' (FORMAT parquet)"
    )
    con.close()
    return cache_path


def load_monitoring(
    con: duckdb.DuckDBPyConnection,
    source_path: str = DEFAULT_MONITORING,
    cache_path: str = DEFAULT_CACHE,
) -> duckdb.DuckDBPyRelation:
    """Ensure the cache, register mon_wide / mon_env / mon_long, return mon_long.

    mon_long is the wide->long contract frame (timestamp, inverter_id, p_ac,
    i_dc, u_dc), built in a single UNPIVOT scan (lazy view). mon_env carries the
    per-timestamp shared tracks with module-temp NaN-clipped (>90 C). The night
    filter (Altitude > 5) is applied by the detection layer.
    """
    build_monitoring_cache(source_path, cache_path)
    con.execute(
        f"CREATE OR REPLACE VIEW mon_wide AS SELECT * FROM read_parquet('{cache_path}')"
    )
    cols = [r[0] for r in con.execute("DESCRIBE SELECT * FROM mon_wide").fetchall()]
    g = resolve_columns(cols)
    tmod = f'"{g["tmod"]}"' if g["tmod"] else "NULL"
    con.execute(
        f"""CREATE OR REPLACE VIEW mon_env AS SELECT
            timestamp AS ts,
            "{g['irr']}" AS irradiation,
            "{g['alt']}" AS altitude,
            "{g['dv']}" AS dv,
            CASE WHEN {tmod} > 90 THEN NULL ELSE {tmod} END AS module_temp
        FROM mon_wide"""
    )
    metric_cols = ",".join(f'"{c}"' for c in g["pac"] + g["idc"] + g["udc"])
    con.execute(
        f"""CREATE OR REPLACE VIEW mon_long AS
        SELECT timestamp AS "timestamp",
               regexp_extract(name, 'INV [0-9.]+') AS inverter_id,
               max(value) FILTER (WHERE name LIKE '%P_AC%') AS p_ac,
               max(value) FILTER (WHERE name LIKE '%I_DC%') AS i_dc,
               max(value) FILTER (WHERE name LIKE '%U_DC%') AS u_dc
        FROM (UNPIVOT mon_wide ON {metric_cols} INTO NAME name VALUE value)
        GROUP BY 1, 2"""
    )
    return con.view("mon_long")


def load_meta(system_overview_path: str = DEFAULT_SYSTEM_OVERVIEW) -> pd.DataFrame:
    """Parse System_Overview.xlsx -> DataFrame[inverter_id, kwp, orientation].

    Header is NOT row 1 (rows 0-1 are 'Project:' / 'Example Plant'); the header
    is the row whose cell == 'Description'. Inverter rows have WR-Type ==
    'Inverter'. PDC (kWp) is summed per 3-part id (split inverters); orientation
    is the resolved 'O'/'W' column (defaults to 'O' if absent).
    """
    import openpyxl

    wb = openpyxl.load_workbook(system_overview_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i = next(
        (
            i
            for i, r in enumerate(rows)
            if any(str(c).strip() == "Description" for c in r if c is not None)
        ),
        None,
    )
    if hdr_i is None:
        raise ValueError("System_Overview: 'Description' header row not found")
    hdr = [("" if c is None else str(c).strip()) for c in rows[hdr_i]]
    desc_c = hdr.index("Description")
    pdc_c = next((i for i, h in enumerate(hdr) if "PDC" in h), None)
    wr_c = next((i for i, h in enumerate(hdr) if "WR-Type" in h), None)
    inv_rows = [
        r
        for r in rows[hdr_i + 1 :]
        if r and wr_c is not None and len(r) > wr_c and str(r[wr_c]).strip() == "Inverter"
    ]
    ori_counts: dict[int, int] = {}
    for r in inv_rows:
        for ci, v in enumerate(r):
            if v is not None and str(v).strip() in ("O", "W"):
                ori_counts[ci] = ori_counts.get(ci, 0) + 1
    ori_c = max(ori_counts, key=lambda c: ori_counts[c]) if ori_counts else None
    kwp: dict[str, float] = {}
    orientation: dict[str, str] = {}
    for r in inv_rows:
        iid = canonical_inverter_id(r[desc_c])
        if not iid:
            continue
        val = r[pdc_c] if (pdc_c is not None and len(r) > pdc_c) else None
        try:
            kwp[iid] = kwp.get(iid, 0.0) + float(val or 0)
        except (TypeError, ValueError):
            pass
        if iid not in orientation and ori_c is not None and len(r) > ori_c and r[ori_c]:
            orientation[iid] = str(r[ori_c]).strip()
    if not orientation:
        print("[ingest] no orientation column resolved; defaulting all inverters to 'O'")
    recs = [
        {"inverter_id": k, "kwp": kwp[k], "orientation": orientation.get(k, "O")}
        for k in kwp
    ]
    return pd.DataFrame(recs)
