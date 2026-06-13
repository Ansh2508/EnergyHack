"""Fleet scan: price every inverter's sustained-zero outage across the whole data
span, using the SAME deterministic sibling counterfactual the hero path uses
(quantify.py, method='sibling_sigma' -- NOT CausalImpact). Backend only: this is
not in the agent decision path and writes only the new outputs/fleet_scan.json.

Reuses, never reimplements:
  - detection + curtailment mask via run_slice1 / outputs/detection_daily.parquet
    (reason=='fault' already excludes curtailment -- the same mask the hero match uses);
  - the deterministic loss via quantify._sibling_sigma_loss (in-memory, exact parity
    with quantify.quantify_loss(prefer_causalimpact=False));
  - the feed-in tariff via quantify._read_tariff (read from feed-in-tarrifs.xlsx,
    never assumed), memoised per (inverter, window).

A "sustained-zero event" is a run of >= MIN_CONSECUTIVE consecutive non-curtailed
fault days with near-zero PR (< ZERO_PR). The near-zero gate excludes the partial
descent / recovery days that border an outage, so there is one event per outage and
no double-counting.
"""

from __future__ import annotations

import datetime as dt
import json
import math
import os
from functools import lru_cache

import pandas as pd
from pydantic import BaseModel

from src import diagnose_events, quantify, run_slice1

DEFAULT_DETECTION = "outputs/detection_daily.parquet"
DEFAULT_TARIFFS = quantify.DEFAULT_TARIFFS
OUT_PATH = "outputs/fleet_scan.json"
HERO_ID = "INV 01.05.029"
MIN_CONSECUTIVE = (
    3  # same sustained-fault gate the hero match uses (hero_match.MIN_CONSECUTIVE)
)
ZERO_PR = 0.10  # "sustained-zero": near-zero output; excludes partial/recovery days


class FleetEvent(BaseModel):
    """One sustained-zero outage priced with the deterministic counterfactual."""

    inverter_id: str
    euros: float
    lost_kwh: float
    n_fault_days: int
    window: dict[str, str]


def load_detection(detection_path: str = DEFAULT_DETECTION) -> pd.DataFrame:
    """Return the fleet-wide inverter-day detections (reason/pr/kwp/insol/...).

    Reuses run_slice1's deliverable. If it is missing, run the Slice-1 pipeline
    once to produce it (detect -> sibling baseline -> curtailment mask).
    """
    if not os.path.exists(detection_path):
        run_slice1.main()
    det = pd.read_parquet(detection_path)
    det["date"] = pd.to_datetime(det["date"]).dt.date
    return det


def _consecutive_runs(days: list[dt.date]) -> list[tuple[dt.date, dt.date]]:
    """Collapse a sorted date list into maximal (start, end) consecutive runs."""
    days = sorted(days)
    if not days:
        return []
    out = []
    start = prev = days[0]
    for d in days[1:]:
        if (d - prev).days == 1:
            prev = d
        else:
            out.append((start, prev))
            start = prev = d
    out.append((start, prev))
    return out


def find_events(
    det: pd.DataFrame,
    min_consecutive: int = MIN_CONSECUTIVE,
    zero_pr: float = ZERO_PR,
) -> list[tuple[str, dt.date, dt.date, int]]:
    """Sustained-zero outages as (inverter_id, start, end, n_days).

    reason=='fault' already excludes curtailment (the same mask the hero path uses);
    the pr < zero_pr gate keeps only genuine near-zero output and drops the partial
    descent / recovery days, so each outage is a single run with no double-counting.
    """
    zero = det[(det["reason"] == "fault") & det["pr"].notna() & (det["pr"] < zero_pr)]
    events: list[tuple[str, dt.date, dt.date, int]] = []
    for iid, g in zero.groupby("inverter_id"):
        for r0, r1 in _consecutive_runs(g["date"].tolist()):
            n = (r1 - r0).days + 1
            if n >= min_consecutive:
                events.append((str(iid), r0, r1, n))
    return events


@lru_cache(maxsize=1)
def _tariff_rows(tariffs_path: str) -> tuple:
    """Load the tariff workbook ONCE (the only workbook read for the whole scan)."""
    import openpyxl

    wb = openpyxl.load_workbook(tariffs_path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    return tuple(tuple(r) for r in sheet.iter_rows(values_only=True))


@lru_cache(maxsize=None)
def _tariff(tariffs_path: str, iid: str, ws_iso: str, we_iso: str) -> float:
    """Feed-in tariff (EUR/kWh) for the window -- read from the file, never assumed.

    Mirrors quantify._read_tariff's window/row/average lookup (and reuses its
    _as_date), over a once-cached copy of the sheet so a fleet-wide scan does not
    re-open the workbook per event.
    """
    ws = pd.Timestamp(ws_iso).date()
    we = pd.Timestamp(we_iso).date()
    rows = _tariff_rows(tariffs_path)
    hdr = rows[1]
    dates = {
        j: quantify._as_date(hdr[j])
        for j in range(1, len(hdr))
        if quantify._as_date(hdr[j])
    }
    cols = [j for j, d in dates.items() if (ws - dt.timedelta(days=7)) <= d <= we]
    row = next((r for r in rows if r and str(r[0]).strip() == iid), None)
    if row is None or not cols:
        raise ValueError(f"tariff not found for {iid} in window")
    vals = []
    for j in cols:
        v = row[j]
        if v in (None, ""):
            continue
        try:
            vals.append(float(str(v).replace(",", ".")))
        except ValueError:
            continue
    if not vals:
        raise ValueError("no numeric tariff values in window")
    ct = sum(vals) / len(vals)
    return ct / 100.0 if ct > 1.0 else ct  # ct/kWh -> EUR/kWh (mirrors quantify)


def price_event(
    inv_rows: pd.DataFrame,
    iid: str,
    r0: dt.date,
    r1: dt.date,
    tariffs_path: str = DEFAULT_TARIFFS,
) -> tuple[float, float, float] | None:
    """Deterministic recoverable loss for one event -> (euros, lost_kwh, tariff).

    Uses quantify._sibling_sigma_loss (expected = sibling_PR * kWp * insolation,
    lost = expected - actual). Returns None when the value is not finite or no tariff
    exists for the window. Euros and kWh are clamped to >= 0.
    """
    pre0 = r0 - dt.timedelta(days=quantify.PRE_DAYS)
    lost, _lo, _hi = quantify._sibling_sigma_loss(inv_rows, r0, r1, pre0)
    if not math.isfinite(lost):
        return None
    try:
        tariff = _tariff(tariffs_path, iid, r0.isoformat(), r1.isoformat())
    except Exception:  # window not covered by the tariff sheet -> skip, do not assume
        return None
    if not math.isfinite(tariff):
        return None
    lost = max(0.0, float(lost))
    euros = max(0.0, lost * tariff)
    return round(euros, 2), round(lost, 1), round(float(tariff), 4)


def scan(
    detection_path: str = DEFAULT_DETECTION,
    tariffs_path: str = DEFAULT_TARIFFS,
) -> dict:
    """Run the fleet scan and return the aggregate (pure compute, no file writes)."""
    det = load_detection(detection_path)
    by_inv = {iid: g for iid, g in det.groupby("inverter_id")}
    events = find_events(det)

    ranked: list[dict] = []
    skipped: list[dict] = []
    for iid, r0, r1, n in events:
        priced = price_event(by_inv[iid], iid, r0, r1, tariffs_path)
        if priced is None:
            skipped.append({"inverter_id": iid, "window": f"{r0}..{r1}"})
            continue
        euros, lost_kwh, _tar = priced
        ranked.append(
            FleetEvent(
                inverter_id=iid,
                euros=euros,
                lost_kwh=lost_kwh,
                n_fault_days=int(n),
                window={"start": r0.isoformat(), "end": r1.isoformat()},
            ).model_dump()
        )

    ranked.sort(key=lambda e: e["euros"], reverse=True)
    euros_list = [e["euros"] for e in ranked]
    kwh_list = [e["lost_kwh"] for e in ranked]
    dmin, dmax = min(det["date"]), max(det["date"])  # REAL span, computed from the data
    hero_events = [e for e in ranked if e["inverter_id"] == HERO_ID]
    hero_check = max(hero_events, key=lambda e: e["euros"]) if hero_events else None

    # snow / weather split: a start date shared by >=5 inverters is a shared cause
    # (weather / section / planned), not a per-inverter recoverable fault. Reuses the
    # diagnostic clustering verbatim; fleet_total_eur and fleet_total_kwh are unchanged.
    clusters = diagnose_events.find_clusters(ranked)
    snow_dates = {c["start"] for c in clusters if c["n"] >= 5}
    for e in ranked:
        e["bucket"] = "weather" if e["window"]["start"] in snow_dates else "isolated"
    iso = [e for e in ranked if e["bucket"] == "isolated"]
    wea = [e for e in ranked if e["bucket"] == "weather"]
    isolated_total_eur = round(sum(e["euros"] for e in iso), 2)
    weather_suppressed_eur = round(sum(e["euros"] for e in wea), 2)
    assert (
        abs(isolated_total_eur + weather_suppressed_eur - round(sum(euros_list), 2))
        < 0.01
    )

    return {
        "data_span": {"start": dmin.isoformat(), "end": dmax.isoformat()},
        "fleet_total_eur": round(sum(euros_list), 2),
        "fleet_total_kwh": round(sum(kwh_list), 1),
        "n_events": len(ranked),
        "n_affected_inverters": len({e["inverter_id"] for e in ranked}),
        "isolated_total_eur": isolated_total_eur,
        "isolated_total_kwh": round(sum(e["lost_kwh"] for e in iso), 1),
        "isolated_events": len(iso),
        "isolated_inverters": len({e["inverter_id"] for e in iso}),
        "weather_suppressed_eur": weather_suppressed_eur,
        "weather_suppressed_events": len(wea),
        "weather_cluster_days": sorted(snow_dates),
        "ranked": ranked,
        "hero_check": hero_check,
        "skipped": skipped,
    }


def write_scan(result: dict, out_path: str = OUT_PATH) -> None:
    """Write the aggregate to outputs/fleet_scan.json (ASCII, LF)."""
    parent = os.path.dirname(out_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(out_path, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(result, fh, indent=2, ensure_ascii=True)
        fh.write("\n")


def _print_summary(r: dict) -> None:
    print("Fleet scan - sustained-zero outages (deterministic sibling counterfactual)")
    print(f"  data span: {r['data_span']['start']} -> {r['data_span']['end']}")
    print(
        f"  fleet total: EUR {r['fleet_total_eur']:,.2f}  ({r['fleet_total_kwh']:,.0f} kWh)"
    )
    print(
        f"  events: {r['n_events']}  |  affected inverters: {r['n_affected_inverters']}"
    )
    hc = r["hero_check"]
    if hc:
        w = hc["window"]
        print(
            f"  hero {hc['inverter_id']}: EUR {hc['euros']:.2f}  "
            f"({w['start']}..{w['end']}, {hc['n_fault_days']}d)"
        )
    print("  top 5 ranked:")
    for e in r["ranked"][:5]:
        w = e["window"]
        print(
            f"    {e['inverter_id']:<14} EUR {e['euros']:>9.2f}  {e['lost_kwh']:>9.1f} kWh  "
            f"{e['n_fault_days']:>2}d  {w['start']}..{w['end']}"
        )
    print(
        f"  isolated faults:    EUR {r['isolated_total_eur']:,.2f} "
        f"({r['isolated_events']} events, {r['isolated_inverters']} inverters)"
    )
    print(
        f"  weather-suppressed: EUR {r['weather_suppressed_eur']:,.2f} "
        f"({r['weather_suppressed_events']} events, {len(r['weather_cluster_days'])} cluster days)"
    )


def main(
    detection_path: str = DEFAULT_DETECTION, tariffs_path: str = DEFAULT_TARIFFS
) -> dict:
    """Run the scan, write outputs/fleet_scan.json, print the summary."""
    r = scan(detection_path, tariffs_path)
    write_scan(r)
    _print_summary(r)
    return r


if __name__ == "__main__":
    main()
