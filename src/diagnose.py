"""Diagnose the cause of an inverter underperformance window -> CauseVerdict.

Each rule is a named, auditable "skill" and the Slice-1 curtailment mask is kept
as a lesson-from-failure guard (SkillRL, arXiv:2602.08234). Tools return
STRUCTURED, inspectable evidence (typed fields + an evidence list), not prose,
so a downstream agent can reason over them (Roy et al. 2024, arXiv:2403.04123).

External methods are verified in docs/RESEARCH_NOTES.md:
  - soiling  : rdtools.soiling.soiling_srr            (Deceglie et al. 2018)
  - clipping : rdtools.filtering.clip_filter('logic') (Perry et al. 2021)
                NOTE: clip_filter returns True=KEEP (not clipped); clipped = ~mask.
"""

from __future__ import annotations

import re

import duckdb
import pandas as pd
from pydantic import BaseModel, Field

from src.ingest import (
    DEFAULT_CACHE,
    DEFAULT_SYSTEM_OVERVIEW,
    STEP_SECONDS,
    canonical_inverter_id,
    load_meta,
)

_DATA = "data/Plant A (start here)"
DEFAULT_ERRORCODES = f"{_DATA}/3. Errorcodes/errorcodes.parquet"
DEFAULT_ERRORCODES_DICT = f"{_DATA}/3. Errorcodes/errorcodes description (important).xlsx"

DEAD_PR_MAX = 0.05  # daily PR below this = effectively no output
NORMAL_INSOL_MIN = 1.0  # kWh/m2/day = a real-sun day (excludes darkness)
UDC_PRESENT_V = 100.0  # U_DC above this during outage => DC side alive => AC fault
IDC_ZERO_A = 1.0  # I_DC below this => no current drawn
MIN_DEAD_DAYS = 3
THERMAL_TMOD_C = 55.0
CURTAILMENT_DOMINATED = 0.5  # >50% of daytime intervals curtailed => not a fault


class CauseVerdict(BaseModel):
    """Structured, inspectable diagnosis (typed fields + evidence list)."""

    inverter_id: str
    primary_cause: str  # DEAD_INVERTER|SOILING|CLIPPING|THERMAL_DERATE|NOT_A_FAULT|UNKNOWN
    side: str | None = None  # 'AC' | 'DC' | None
    confidence: float = Field(ge=0.0, le=1.0)
    evidence: list[str]
    errorcode_corroboration: str | None = None


def diagnose(
    inverter_id: str,
    window_start: str,
    window_end: str,
    *,
    cache_path: str = DEFAULT_CACHE,
    errorcodes_path: str = DEFAULT_ERRORCODES,
    errorcodes_dict_path: str = DEFAULT_ERRORCODES_DICT,
    system_overview_path: str = DEFAULT_SYSTEM_OVERVIEW,
) -> CauseVerdict:
    """Classify the cause for inverter_id over [window_start, window_end]."""
    iid = canonical_inverter_id(inverter_id) or str(inverter_id)
    ws = pd.Timestamp(window_start).date()
    we = pd.Timestamp(window_end).date()
    meta = load_meta(system_overview_path)
    mrow = meta[meta["inverter_id"] == iid]
    kwp = float(mrow["kwp"].iloc[0]) if not mrow.empty else 0.0
    if kwp <= 0:
        raise ValueError(f"no kWp for {iid} in System_Overview")
    df = _inverter_daily(iid, ws, we, cache_path, kwp)
    ec = _errorcode_corroboration(iid, ws, we, errorcodes_path, errorcodes_dict_path)

    # priority 0 - lesson-from-failure guard: never score a market throttle as a fault
    if _skill_curtailment_guard(iid, ws, we, cache_path):
        return CauseVerdict(
            inverter_id=iid,
            primary_cause="NOT_A_FAULT",
            confidence=0.9,
            evidence=["window is curtailment-dominated (DV<100): grid/market throttle, not a fault"],
            errorcode_corroboration=ec,
        )

    # priority 1 - dead inverter / AC outage, with DC-side disambiguation
    dead = _skill_dead_inverter(df, ws, we)
    if dead["hit"]:
        conf = 0.7 + (0.15 if dead["side"] else 0.0) + (0.1 if ec else 0.0)
        conf += min(0.04, dead["consecutive_days"] / 300.0)
        return CauseVerdict(
            inverter_id=iid,
            primary_cause="DEAD_INVERTER",
            side=dead["side"],
            confidence=round(min(conf, 0.99), 2),
            evidence=dead["evidence"],
            errorcode_corroboration=ec,
        )

    win = df[(df["date"] >= ws) & (df["date"] <= we)].dropna(subset=["pr", "insol"])

    # priority 2 - soiling (gradual decline + recovery)
    try:
        s = _skill_soiling(_as_series(win, "pr"), _as_series(win, "insol"))
    except Exception as exc:  # soiling_srr can reject short/degenerate series
        s = {"hit": False, "evidence": [f"soiling_srr not applicable: {exc}"]}
    if s["hit"]:
        return CauseVerdict(
            inverter_id=iid,
            primary_cause="SOILING",
            confidence=0.6,
            evidence=s["evidence"],
            errorcode_corroboration=ec,
        )

    # priority 3 - clipping (flat at the TOP near rated)
    try:
        c = _skill_clipping(_inverter_power_5min(iid, ws, we, cache_path), kwp)
    except Exception as exc:
        c = {"hit": False, "evidence": [f"clip_filter not applicable: {exc}"]}
    if c["hit"]:
        return CauseVerdict(
            inverter_id=iid,
            primary_cause="CLIPPING",
            confidence=0.6,
            evidence=c["evidence"],
            errorcode_corroboration=ec,
        )

    # priority 4 - thermal derate
    t = _skill_thermal(win)
    if t["hit"]:
        return CauseVerdict(
            inverter_id=iid,
            primary_cause="THERMAL_DERATE",
            confidence=0.55,
            evidence=t["evidence"],
            errorcode_corroboration=ec,
        )

    # priority 5 - unknown (report the signature honestly, do not force a label)
    sig = f"window mean PR={win['pr'].mean():.2f}" if not win.empty else "no daytime data"
    return CauseVerdict(
        inverter_id=iid,
        primary_cause="UNKNOWN",
        confidence=0.3,
        evidence=["no rule matched; signature reported honestly", sig],
        errorcode_corroboration=ec,
    )


# --------------------------------------------------------------------------- #
# named diagnostic skills
# --------------------------------------------------------------------------- #
def _skill_dead_inverter(df: pd.DataFrame, ws, we) -> dict:
    """P_AC ~0 for >= MIN_DEAD_DAYS consecutive sunny days; split AC vs DC side."""
    w = df[(df["date"] >= ws) & (df["date"] <= we)].copy()
    if w.empty:
        return {"hit": False, "evidence": ["no daytime data in window"]}
    dead_dates = w.loc[w["pr"].fillna(0) < DEAD_PR_MAX, "date"].tolist()
    runs = _consecutive_runs(dead_dates)
    best = max(runs, key=lambda r: (r[1] - r[0]).days, default=None)
    if best is None:
        return {"hit": False, "evidence": ["no near-zero-PR run"]}
    run = w[(w["date"] >= best[0]) & (w["date"] <= best[1])]
    ndays = (best[1] - best[0]).days + 1
    if ndays < MIN_DEAD_DAYS or float(run["insol"].median()) <= NORMAL_INSOL_MIN:
        return {"hit": False, "evidence": [f"near-zero run {ndays}d but not on sunny days"]}
    udc = float(run["udc"].mean())
    idc = float(run["idc"].mean())
    if udc > UDC_PRESENT_V and idc < IDC_ZERO_A:
        side = "AC"
        side_txt = (
            f"U_DC present (~{udc:.0f} V, open-circuit) with I_DC~0 ({idc:.2f} A): "
            "panels healthy, inverter (AC side) failed"
        )
    elif udc <= UDC_PRESENT_V and idc < IDC_ZERO_A:
        side = "DC"
        side_txt = (
            f"U_DC~0 ({udc:.0f} V) and I_DC~0 ({idc:.2f} A): "
            "no DC reaching the inverter (string/DC side failure)"
        )
    else:
        side = None
        side_txt = f"ambiguous DC side (U_DC {udc:.0f} V, I_DC {idc:.2f} A)"
    ev = [
        f"P_AC ~0 for {ndays} consecutive days {best[0]}..{best[1]} "
        f"while insolation normal (median {float(run['insol'].median()):.1f} kWh/m2)",
        side_txt,
    ]
    return {"hit": True, "side": side, "consecutive_days": ndays, "evidence": ev}


def _skill_soiling(pr_daily: pd.Series, insol_daily: pd.Series,
                   confidence_level: float = 95.0, reps: int = 200) -> dict:
    """rdtools soiling_srr (Deceglie 2018); SOILING only on gradual, non-zero decline."""
    if float((pr_daily < 0.1).mean()) > 0.2:
        return {"hit": False, "soiling_ratio": None, "ci": None,
                "evidence": ["near-zero days present -> outage, not gradual soiling"]}
    import warnings

    from rdtools.soiling import soiling_srr

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        sr, ci, _info = soiling_srr(
            pr_daily, insol_daily, confidence_level=confidence_level, reps=reps
        )
    hit = 0.5 < sr < 0.975
    ev = [f"soiling_srr P50 ratio {sr:.3f} (95% CI {ci[0]:.3f}-{ci[1]:.3f})"]
    return {"hit": bool(hit), "soiling_ratio": float(sr),
            "ci": [float(ci[0]), float(ci[1])], "evidence": ev}


def _skill_clipping(power_ac: pd.Series, rated: float) -> dict:
    """rdtools clip_filter('logic', Perry 2021). Clipped = flat at the TOP, not at zero."""
    from rdtools.filtering import clip_filter

    p = power_ac.dropna()
    keep = clip_filter(p, model="logic")  # True = keep (not clipped)
    clipped = ~keep
    n = int(clipped.sum())
    if n == 0 or not clipped.any():
        return {"hit": False, "evidence": ["clip_filter: no clipping intervals"]}
    level = float(p[clipped].median())
    hit = level > 0.5 * rated  # near rated => clipping; near zero => dead, not clipping
    return {"hit": bool(hit),
            "evidence": [f"clip_filter flagged {n} intervals at median {level:.1f} kW "
                         f"(rated ~{rated:.1f} kW)"]}


def _skill_thermal(win: pd.DataFrame) -> dict:
    """PR dips on hot days (module temp > THERMAL_TMOD_C) relative to cool days."""
    w = win.dropna(subset=["pr", "tmod"])
    hot = w[w["tmod"] > THERMAL_TMOD_C]
    cool = w[w["tmod"] <= THERMAL_TMOD_C]
    if len(hot) < 2 or len(cool) < 2:
        return {"hit": False, "evidence": ["insufficient hot/cool days for thermal test"]}
    drop = float(cool["pr"].mean() - hot["pr"].mean())
    return {"hit": bool(drop > 0.05),
            "evidence": [f"PR on hot days (Tmod>{THERMAL_TMOD_C}C) is {drop:+.3f} below cool days"]}


def _skill_curtailment_guard(inverter_id: str, ws, we, cache_path: str) -> bool:
    """Lesson-from-failure: True if the window is curtailment-dominated (DV<100)."""
    con = duckdb.connect()
    try:
        cols = [r[0] for r in con.execute(
            f"DESCRIBE SELECT * FROM read_parquet('{cache_path}')").fetchall()]
        dv = next((c for c in cols if re.search(r"/ DV", c)), None)
        alt = next((c for c in cols if re.search(r"Altitude", c)), None)
        if not dv or not alt:
            return False
        frac = con.execute(
            f"""SELECT avg(CASE WHEN "{dv}" < 100 THEN 1.0 ELSE 0.0 END)
                FROM read_parquet('{cache_path}')
                WHERE "{alt}" > 5 AND CAST(timestamp AS DATE) BETWEEN '{ws}' AND '{we}'"""
        ).fetchone()[0]
    finally:
        con.close()
    return bool(frac is not None and frac > CURTAILMENT_DOMINATED)


# --------------------------------------------------------------------------- #
# data access + helpers
# --------------------------------------------------------------------------- #
def _inverter_daily(inverter_id: str, ws, we, cache_path: str, kwp: float,
                    context_days: int = 14) -> pd.DataFrame:
    """Daytime daily series for one inverter (date, daily_kwh, idc, udc, tmod, insol, pr)."""
    con = duckdb.connect()
    try:
        cols = [r[0] for r in con.execute(
            f"DESCRIBE SELECT * FROM read_parquet('{cache_path}')").fetchall()]
        eid = re.escape(inverter_id)
        pac = _find(cols, eid + r" / P_AC")
        idc = _find(cols, eid + r" / I_DC")
        udc = _find(cols, eid + r" / U_DC")
        irr = _find(cols, r"Irradiation")
        alt = _find(cols, r"Altitude")
        tmod = _find(cols, r"/ Module")
        if not pac:
            raise ValueError(f"no P_AC column for {inverter_id}")
        start_ctx = (pd.Timestamp(ws) - pd.Timedelta(days=context_days)).date()
        df = con.execute(
            f"""WITH d AS (
                  SELECT timestamp ts, "{pac}" pac, "{idc}" idc, "{udc}" udc, "{irr}" irr
                       {(', "' + tmod + '" tmod') if tmod else ', NULL tmod'}
                  FROM read_parquet('{cache_path}')
                  WHERE "{alt}" > 5
                    AND CAST(timestamp AS DATE) BETWEEN '{start_ctx}' AND '{we}')
                SELECT CAST(ts AS DATE) date,
                       sum(pac) * {STEP_SECONDS} / 3600.0 AS daily_kwh,
                       avg(idc) AS idc, avg(udc) AS udc, avg(tmod) AS tmod,
                       sum(irr) * {STEP_SECONDS} / 3600 / 1000.0 AS insol
                FROM d GROUP BY 1 ORDER BY 1"""
        ).df()
    finally:
        con.close()
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["pr"] = df["daily_kwh"] / (kwp * df["insol"].replace(0, pd.NA))
    return df


def _inverter_power_5min(inverter_id: str, ws, we, cache_path: str) -> pd.Series:
    """Daytime 5-min P_AC series for one inverter (for the clipping skill)."""
    con = duckdb.connect()
    try:
        cols = [r[0] for r in con.execute(
            f"DESCRIBE SELECT * FROM read_parquet('{cache_path}')").fetchall()]
        eid = re.escape(inverter_id)
        pac = _find(cols, eid + r" / P_AC")
        alt = _find(cols, r"Altitude")
        df = con.execute(
            f"""SELECT timestamp ts, "{pac}" p FROM read_parquet('{cache_path}')
                WHERE "{alt}" > 5 AND CAST(timestamp AS DATE) BETWEEN '{ws}' AND '{we}'
                ORDER BY ts"""
        ).df()
    finally:
        con.close()
    return pd.Series(df["p"].to_numpy(), index=pd.to_datetime(df["ts"]))


def _errorcode_corroboration(inverter_id: str, ws, we, ec_path: str, dict_path: str) -> str | None:
    """Top non-zero Refu error code for this inverter in the window, mapped to German text."""
    con = duckdb.connect()
    try:
        cols = [r[0] for r in con.execute(
            f"DESCRIBE SELECT * FROM read_parquet('{ec_path}')").fetchall()]
        eid = re.escape(inverter_id)
        errc = _find(cols, eid + r" / Error")
        if not errc:
            return None
        row = con.execute(
            f"""SELECT "{errc}" code, count(*) n FROM read_parquet('{ec_path}')
                WHERE "{errc}" IS NOT NULL AND "{errc}" != 0
                  AND CAST(strptime(timestamp, '%Y.%m.%d %H:%M') AS DATE)
                      BETWEEN '{ws}' AND '{we}'
                GROUP BY 1 ORDER BY n DESC LIMIT 1"""
        ).fetchone()
    finally:
        con.close()
    if not row:
        return None
    code, n = int(row[0]), int(row[1])
    text = _load_errorcode_dict(dict_path).get(code, "unknown code")
    return f"{code} ({n}x in window): {text}"


def _load_errorcode_dict(dict_path: str) -> dict[int, str]:
    """Map Refu Dezimal code -> German fault description from the dictionary xlsx."""
    import openpyxl

    wb = openpyxl.load_workbook(dict_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[int, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4 or row[2] is None:
            continue
        try:
            out[int(row[2])] = str(row[3]) if row[3] is not None else ""
        except (TypeError, ValueError):
            continue
    return out


def _as_series(win: pd.DataFrame, col: str) -> pd.Series:
    return pd.Series(win[col].to_numpy(), index=pd.to_datetime(win["date"]))


def _find(cols: list[str], pattern: str) -> str | None:
    """First column matching a regex, else None (never hardcode column names)."""
    return next((c for c in cols if re.search(pattern, c)), None)


def _consecutive_runs(dates: list) -> list:
    dates = sorted(dates)
    if not dates:
        return []
    runs = []
    start = prev = dates[0]
    for d in dates[1:]:
        if (d - prev).days == 1:
            prev = d
        else:
            runs.append((start, prev))
            start = prev = d
    runs.append((start, prev))
    return runs


if __name__ == "__main__":
    verdict = diagnose("INV 01.05.029", "2019-05-24", "2019-06-16")
    print("CauseVerdict for INV 01.05.029:")
    for key, val in verdict.model_dump().items():
        print(f"  {key}: {val}")
