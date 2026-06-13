"""Quantify lost energy + euros for a fault window -> LossEstimate.

Counterfactual = Brodersen et al. 2015 (Bayesian structural time series). The
primary path uses the maintained port tfcausalimpact:
    ci = CausalImpact(data, pre_period, post_period, model_args={'fit_method':'vi'})
and reads the cumulative effect + 95% CI from the verified attribute path
    ci.summary_data.loc['abs_effect'|'abs_effect_lower'|'abs_effect_upper', 'cumulative']
(see docs/RESEARCH_NOTES.md). When TensorFlow is unavailable or the VI fit will
not converge, a labelled sibling-counterfactual fallback runs instead
(method='sibling_sigma'): expected = sibling_median_PR * kWp * insolation,
lost = expected - actual, with a 95% band of +/- 1.96 * sigma * sqrt(n) from the
pre-period daily residual standard deviation. The method is always reported
honestly. The feed-in tariff is READ from feed-in-tarrifs.xlsx, never assumed.
"""

from __future__ import annotations

import datetime as dt
import math

import openpyxl
import pandas as pd
from pydantic import BaseModel

from src.ingest import canonical_inverter_id

_DATA = "data/Plant A (start here)"
DEFAULT_DETECTION = "outputs/detection_daily.parquet"
DEFAULT_TARIFFS = f"{_DATA}/2. Additional Data/feed-in-tarrifs.xlsx"
PRE_DAYS = 60
N_CONTROLS = 6
Z95 = 1.96


class LossEstimate(BaseModel):
    """Lost energy + euros with a 95% interval and the method that produced it."""

    inverter_id: str
    lost_kwh: float
    lost_kwh_ci_low: float
    lost_kwh_ci_high: float
    tariff_eur_per_kwh: float
    euros_lost: float
    euros_ci_low: float
    euros_ci_high: float
    method: str  # 'causalimpact' | 'sibling_sigma'
    pre_period: str
    post_period: str


def quantify_loss(
    inverter_id: str,
    window_start: str,
    window_end: str,
    *,
    detection_path: str = DEFAULT_DETECTION,
    tariffs_path: str = DEFAULT_TARIFFS,
    prefer_causalimpact: bool = True,
) -> LossEstimate:
    """Estimate lost kWh + euros for inverter_id over [window_start, window_end]."""
    iid = canonical_inverter_id(inverter_id) or str(inverter_id)
    ws = pd.Timestamp(window_start).date()
    we = pd.Timestamp(window_end).date()
    pre0 = ws - dt.timedelta(days=PRE_DAYS)
    det = pd.read_parquet(detection_path)
    det["date"] = pd.to_datetime(det["date"]).dt.date
    hero = det[det["inverter_id"] == iid].copy()
    if hero.empty:
        raise ValueError(f"no detection rows for {iid}")

    method: str | None = None
    lost = ci_lo = ci_hi = 0.0
    if prefer_causalimpact:
        try:
            lost, ci_lo, ci_hi = _causalimpact_loss(det, iid, ws, we, pre0)
            method = "causalimpact"
        except Exception:  # TF missing or VI did not converge -> documented fallback
            method = None
    if method is None:
        lost, ci_lo, ci_hi = _sibling_sigma_loss(hero, ws, we, pre0)
        method = "sibling_sigma"

    tariff = _read_tariff(tariffs_path, iid, ws, we)
    return LossEstimate(
        inverter_id=iid,
        lost_kwh=round(lost, 1),
        lost_kwh_ci_low=round(ci_lo, 1),
        lost_kwh_ci_high=round(ci_hi, 1),
        tariff_eur_per_kwh=round(tariff, 4),
        euros_lost=round(lost * tariff, 2),
        euros_ci_low=round(ci_lo * tariff, 2),
        euros_ci_high=round(ci_hi * tariff, 2),
        method=method,
        pre_period=f"{pre0}..{ws - dt.timedelta(days=1)}",
        post_period=f"{ws}..{we}",
    )


def _sibling_sigma_loss(hero: pd.DataFrame, ws, we, pre0):
    """Sibling-counterfactual: expected = sibling_PR * kWp * insolation."""
    h = hero.copy()
    h["expected"] = h["sibling_pr"] * h["kwp"] * h["insol_kwhm2"]
    pre = h[(h["date"] >= pre0) & (h["date"] < ws)].dropna(subset=["expected", "daily_kwh"])
    post = h[(h["date"] >= ws) & (h["date"] <= we)].dropna(subset=["expected", "daily_kwh"])
    sigma = float((pre["daily_kwh"] - pre["expected"]).std())
    lost = float((post["expected"] - post["daily_kwh"]).sum())
    half = Z95 * sigma * math.sqrt(max(len(post), 1))
    return lost, lost - half, lost + half


def _causalimpact_loss(det: pd.DataFrame, iid, ws, we, pre0):
    """BSTS counterfactual via tfcausalimpact; verified summary_data attribute path."""
    from causalimpact import CausalImpact

    frame, pre_period, post_period = _ci_frame(det, iid, ws, we, pre0)
    ci = CausalImpact(frame, pre_period, post_period, model_args={"fit_method": "vi"})
    sd = ci.summary_data
    lost = abs(float(sd.loc["abs_effect", "cumulative"]))
    low = abs(float(sd.loc["abs_effect_lower", "cumulative"]))
    high = abs(float(sd.loc["abs_effect_upper", "cumulative"]))
    return lost, min(low, high), max(low, high)


def _ci_frame(det: pd.DataFrame, iid, ws, we, pre0):
    """Response = hero daily kWh; controls = best-correlated healthy siblings."""
    span = det[(det["date"] >= pre0) & (det["date"] <= we)]
    hero = span[span["inverter_id"] == iid].set_index("date")["daily_kwh"].sort_index()
    flagged = set(span[span["reason"] == "fault"]["inverter_id"].unique())
    healthy = [i for i in span["inverter_id"].unique() if i not in flagged and i != iid]
    piv = (
        span[span["inverter_id"].isin(healthy)]
        .pivot(index="date", columns="inverter_id", values="daily_kwh")
        .sort_index()
    )
    pre_idx = [d for d in hero.index if d < ws]
    corr = piv.loc[pre_idx].corrwith(hero.loc[pre_idx]).dropna().sort_values(ascending=False)
    top = corr.head(N_CONTROLS).index.tolist()
    frame = pd.concat([hero.rename("y"), piv[top]], axis=1).dropna()
    frame.index = pd.to_datetime(frame.index)
    return frame, [str(pre0), str(ws - dt.timedelta(days=1))], [str(ws), str(we)]


def _read_tariff(tariffs_path: str, iid: str, ws, we) -> float:
    """EUR/kWh for the window weeks, read from feed-in-tarrifs.xlsx (never assumed)."""
    wb = openpyxl.load_workbook(tariffs_path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    hdr = rows[1]
    dates = {j: _as_date(hdr[j]) for j in range(1, len(hdr)) if _as_date(hdr[j])}
    cols = [j for j, d in dates.items() if (ws - dt.timedelta(days=7)) <= d <= we]
    hero_row = next((r for r in rows if r and str(r[0]).strip() == iid), None)
    if hero_row is None or not cols:
        raise ValueError(f"tariff not found for {iid} in window")
    vals = []
    for j in cols:
        v = hero_row[j]
        if v in (None, ""):
            continue
        try:
            vals.append(float(str(v).replace(",", ".")))
        except ValueError:
            continue
    if not vals:
        raise ValueError("no numeric tariff values in window")
    ct = sum(vals) / len(vals)
    return ct / 100.0 if ct > 1.0 else ct  # stored as ct/kWh -> EUR/kWh


def _as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y.%m.%d"):
            try:
                return dt.datetime.strptime(v[:10], fmt).date()
            except ValueError:
                continue
    return None


if __name__ == "__main__":
    est = quantify_loss("INV 01.05.029", "2019-05-24", "2019-06-16")
    print("LossEstimate for INV 01.05.029:")
    for key, val in est.model_dump().items():
        print(f"  {key}: {val}")
